"""
Excel report generator for inventory forecasting results.
Creates formatted Excel files with predictions and analysis.
"""
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from typing import Dict, List, Optional

from utils.logger import setup_logger

logger = setup_logger("excel_reporter")


class ExcelReporter:
    """Generate formatted Excel reports for inventory forecasts."""
    
    def __init__(self, output_dir: str = "outputs"):
        """
        Initialize Excel reporter.
        
        Args:
            output_dir: Directory to save reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logger
        
        # Define color schemes for urgency levels
        self.colors = {
            'critical': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),  # Red
            'warning': PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid'),   # Yellow
            'normal': PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid'),    # Green
            'header': PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid'),    # Blue
            'subheader': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')  # Gray
        }
        
    def create_forecast_report(
        self,
        forecasts: pd.DataFrame,
        historical_data: Optional[pd.DataFrame] = None,
        model_metrics: Optional[Dict] = None,
        feature_importance: Optional[pd.DataFrame] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        Create comprehensive forecast report.
        
        Args:
            forecasts: DataFrame with columns [part_id, date, forecast, lower_bound, upper_bound]
            historical_data: Historical consumption data
            model_metrics: Dictionary of model performance metrics
            feature_importance: DataFrame with feature importance scores
            filename: Custom filename (default: forecast_report_YYYYMMDD.xlsx)
            
        Returns:
            Path to saved Excel file
        """
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add sheets
        self._create_summary_sheet(wb, forecasts, model_metrics)
        self._create_critical_orders_sheet(wb, forecasts)
        self._create_forecast_details_sheet(wb, forecasts)
        
        if historical_data is not None:
            self._create_historical_analysis_sheet(wb, historical_data)
            
        if feature_importance is not None:
            self._create_feature_importance_sheet(wb, feature_importance)
            
        if model_metrics is not None:
            self._create_metrics_sheet(wb, model_metrics)
            
        # Save file
        if filename is None:
            filename = f"forecast_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        self.logger.info(f"Excel report saved to: {filepath}")
        return str(filepath)
    
    def _create_summary_sheet(self, wb: Workbook, forecasts: pd.DataFrame, metrics: Optional[Dict]):
        """Create summary dashboard sheet."""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Inventory Forecast Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Report metadata
        ws['A3'] = "Report Generated:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws['A4'] = "Forecast Horizon:"
        ws['B4'] = f"{forecasts['date'].nunique()} days"
        ws['A5'] = "Total Parts:"
        ws['B5'] = forecasts['part_id'].nunique()
        
        # Critical metrics
        ws['A7'] = "Critical Metrics"
        ws['A7'].font = Font(size=12, bold=True)
        ws['A7'].fill = self.colors['header']
        ws.merge_cells('A7:D7')
        
        # Calculate critical metrics
        if 'urgency' in forecasts.columns:
            critical_count = (forecasts['urgency'] == 'critical').sum()
            warning_count = (forecasts['urgency'] == 'warning').sum()
        else:
            critical_count = 0
            warning_count = 0
            
        ws['A8'] = "Critical Orders (Need within 2 days):"
        ws['B8'] = critical_count
        ws['B8'].fill = self.colors['critical'] if critical_count > 0 else self.colors['normal']
        
        ws['A9'] = "Warning Orders (Need within 7 days):"
        ws['B9'] = warning_count
        ws['B9'].fill = self.colors['warning'] if warning_count > 0 else self.colors['normal']
        
        # Model performance
        if metrics:
            ws['A11'] = "Model Performance"
            ws['A11'].font = Font(size=12, bold=True)
            ws['A11'].fill = self.colors['header']
            ws.merge_cells('A11:D11')
            
            row = 12
            for metric, value in metrics.items():
                ws[f'A{row}'] = metric.replace('_', ' ').title()
                ws[f'B{row}'] = f"{value:.3f}" if isinstance(value, float) else str(value)
                row += 1
                
        # Format column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
    def _create_critical_orders_sheet(self, wb: Workbook, forecasts: pd.DataFrame):
        """Create sheet with critical orders requiring immediate attention."""
        ws = wb.create_sheet("Critical Orders")
        
        # Filter for critical items
        critical_df = forecasts[forecasts.get('urgency', '') == 'critical'].copy()
        
        if critical_df.empty:
            ws['A1'] = "No Critical Orders"
            ws['A1'].font = Font(size=14, bold=True, color='006600')
            ws['A3'] = "All parts have sufficient inventory for the forecast period."
            return
            
        # Sort by urgency
        critical_df = critical_df.sort_values(['date', 'forecast'], ascending=[True, False])
        
        # Add headers
        headers = ['Part ID', 'Stock Code', 'Date Needed', 'Forecast Qty', 
                  'Min Qty', 'Max Qty', 'Lead Time', 'Order By']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = self.colors['header']
            cell.alignment = Alignment(horizontal='center')
            
        # Add data
        row = 2
        for _, item in critical_df.iterrows():
            ws.cell(row=row, column=1, value=item.get('part_id', ''))
            ws.cell(row=row, column=2, value=item.get('stock_code', ''))
            ws.cell(row=row, column=3, value=item.get('date', '').strftime('%Y-%m-%d') 
                   if pd.notna(item.get('date')) else '')
            ws.cell(row=row, column=4, value=round(item.get('forecast', 0)))
            ws.cell(row=row, column=5, value=round(item.get('lower_bound', 0)) 
                   if 'lower_bound' in item else '')
            ws.cell(row=row, column=6, value=round(item.get('upper_bound', 0))
                   if 'upper_bound' in item else '')
            ws.cell(row=row, column=7, value=item.get('lead_time_days', 7))
            
            # Calculate order by date
            order_by = item.get('date') - timedelta(days=item.get('lead_time_days', 7))
            ws.cell(row=row, column=8, value=order_by.strftime('%Y-%m-%d')
                   if pd.notna(order_by) else '')
            
            # Apply urgency coloring to entire row
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = self.colors['critical']
                
            row += 1
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_forecast_details_sheet(self, wb: Workbook, forecasts: pd.DataFrame):
        """Create detailed forecast sheet with all predictions."""
        ws = wb.create_sheet("Forecast Details")
        
        # Prepare data
        detail_df = forecasts.copy()
        
        # Sort by part and date
        detail_df = detail_df.sort_values(['part_id', 'date'])
        
        # Add headers
        headers = list(detail_df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = self.colors['header']
            cell.alignment = Alignment(horizontal='center')
            
        # Add data
        for r_idx, row in enumerate(dataframe_to_rows(detail_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format dates
                if isinstance(value, datetime):
                    cell.value = value.strftime('%Y-%m-%d')
                    
                # Format numbers
                elif isinstance(value, (int, float)):
                    if headers[c_idx-1] in ['forecast', 'lower_bound', 'upper_bound']:
                        cell.value = round(value, 2)
                        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze top row
        ws.freeze_panes = 'A2'
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_historical_analysis_sheet(self, wb: Workbook, historical_data: pd.DataFrame):
        """Create sheet with historical demand analysis."""
        ws = wb.create_sheet("Historical Analysis")
        
        # Calculate summary statistics
        summary = historical_data.groupby('part_id').agg({
            'consumption': ['mean', 'std', 'min', 'max', 'count'],
            'date': ['min', 'max']
        }).round(2)
        
        # Flatten column names
        summary.columns = ['_'.join(col).strip() for col in summary.columns]
        summary.reset_index(inplace=True)
        
        # Add headers
        headers = ['Part ID', 'Avg Consumption', 'Std Dev', 'Min', 'Max', 
                  'Data Points', 'First Date', 'Last Date']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = self.colors['header']
            
        # Add data
        for r_idx, row in enumerate(summary.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
                
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_feature_importance_sheet(self, wb: Workbook, feature_importance: pd.DataFrame):
        """Create sheet with feature importance analysis."""
        ws = wb.create_sheet("Feature Importance")
        
        # Sort by importance
        feature_importance = feature_importance.sort_values('importance', ascending=False).head(20)
        
        # Add title
        ws['A1'] = "Top 20 Most Important Features"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Add headers
        ws['A3'] = "Rank"
        ws['B3'] = "Feature"
        ws['C3'] = "Importance Score"
        
        for col in ['A3', 'B3', 'C3']:
            ws[col].font = Font(bold=True, color='FFFFFF')
            ws[col].fill = self.colors['header']
            
        # Add data
        for idx, (_, row) in enumerate(feature_importance.iterrows(), 4):
            ws[f'A{idx}'] = idx - 3
            ws[f'B{idx}'] = row['feature']
            ws[f'C{idx}'] = round(row['importance'], 4)
            
        # Create bar chart
        chart = LineChart()
        chart.title = "Feature Importance"
        chart.x_axis.title = "Features"
        chart.y_axis.title = "Importance Score"
        
        # Set data
        data = Reference(ws, min_col=3, min_row=3, max_row=min(23, len(feature_importance)+3))
        categories = Reference(ws, min_col=2, min_row=4, max_row=min(23, len(feature_importance)+3))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, "E3")
        
        # Format columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        
    def _create_metrics_sheet(self, wb: Workbook, metrics: Dict):
        """Create sheet with detailed model metrics."""
        ws = wb.create_sheet("Model Metrics")
        
        # Add title
        ws['A1'] = "Model Performance Metrics"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Add headers
        ws['A3'] = "Metric"
        ws['B3'] = "Value"
        ws['C3'] = "Description"
        
        for col in ['A3', 'B3', 'C3']:
            ws[col].font = Font(bold=True, color='FFFFFF')
            ws[col].fill = self.colors['header']
            
        # Metric descriptions
        descriptions = {
            'mae': 'Mean Absolute Error - Average prediction error',
            'rmse': 'Root Mean Square Error - Penalizes large errors',
            'mape': 'Mean Absolute Percentage Error - Error as percentage',
            'zero_accuracy': 'Accuracy in predicting zero demand days',
            'train_mae': 'MAE on training data',
            'val_mae': 'MAE on validation data',
            'coverage_95': '95% prediction interval coverage'
        }
        
        # Add metrics
        row = 4
        for metric, value in metrics.items():
            ws[f'A{row}'] = metric.replace('_', ' ').title()
            
            if isinstance(value, float):
                ws[f'B{row}'] = round(value, 4)
            else:
                ws[f'B{row}'] = str(value)
                
            ws[f'C{row}'] = descriptions.get(metric, '')
            row += 1
            
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50


def create_quick_report(
    forecasts: pd.DataFrame,
    output_dir: str = "outputs"
) -> str:
    """
    Quick function to create a basic forecast report.
    
    Args:
        forecasts: DataFrame with forecast data
        output_dir: Directory to save report
        
    Returns:
        Path to saved file
    """
    reporter = ExcelReporter(output_dir)
    return reporter.create_forecast_report(forecasts)